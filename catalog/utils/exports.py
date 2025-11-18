from __future__ import annotations

from datetime import date
from typing import List, Optional, Sequence

from django.db.models import Q, Count, Avg, Max, QuerySet, Prefetch
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from catalog.models import Book, Loan


def _parse_int(val: Optional[str]) -> Optional[int]:
    try:
        return int(val) if val is not None and val != "" else None
    except ValueError:
        return None


def _parse_date(val: Optional[str]) -> Optional[date]:
    try:
        return date.fromisoformat(val) if val else None
    except ValueError:
        return None


def annotate_book_metrics(
    qs: QuerySet[Book], include_items: bool = False
) -> QuerySet[Book]:
    """Annotate a Book queryset with counts and engagement metrics and prefetch related data."""
    qs = (
        qs.select_related("publisher")
        .prefetch_related("authors", "categories", "items")
        .annotate(
            total_items=Count("items", distinct=True),
            available_items=Count(
                "items", filter=Q(items__status="AVAILABLE"), distinct=True
            ),
            reserved_items=Count(
                "items", filter=Q(items__status="RESERVED"), distinct=True
            ),
            loaned_items=Count(
                "items", filter=Q(items__status="LOANED"), distinct=True
            ),
            lost_items=Count(
                "items", filter=Q(items__status="LOST"), distinct=True
            ),
            damaged_items=Count(
                "items", filter=Q(items__status="DAMAGED"), distinct=True
            ),
            total_loans=Count("requested_items__loans", distinct=True),
            last_loan_date=Max("requested_items__loans__approved_from"),
            avg_rating=Avg("ratings__rating"),
            total_favorites=Count("favorited_by", distinct=True),
            authors_count=Count("authors", distinct=True),
            categories_count=Count("categories", distinct=True),
        )
    )
    if include_items:
        qs = qs.prefetch_related(
            Prefetch("items__loans", queryset=Loan.objects.all())
        )
    return qs


def build_book_queryset(params, include_items: bool = False) -> QuerySet[Book]:
    """Build filtered/annotated Book queryset based on request GET-like mapping.

    Supports filters: q, category_id, author_id, publisher_id, publish_year_from,
    publish_year_to, language, item_status, created_from, created_to, sort.
    Adds annotations for item counts by status and engagement metrics.
    """
    qs = annotate_book_metrics(Book.objects.all(), include_items=include_items)

    # Support both our custom params and Django admin changelist params
    # Admin equivalents (examples):
    # - categories__id__exact, authors__id__exact, publisher__id__exact
    # - publish_year (exact), publish_year__gte, publish_year__lte
    # - language_code, language_code__exact
    # - created_at__date__gte, created_at__date__lte

    q = (params.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(title__icontains=q)
            | Q(description__icontains=q)
            | Q(isbn13__icontains=q)
        )

    category_id = _parse_int(
        params.get("category_id")
        or params.get("categories__id__exact")
        or params.get("category__id__exact")
    )
    if category_id is not None:
        qs = qs.filter(categories__id=category_id)

    author_id = _parse_int(
        params.get("author_id")
        or params.get("authors__id__exact")
        or params.get("author__id__exact")
    )
    if author_id is not None:
        qs = qs.filter(authors__id=author_id)

    publisher_id = _parse_int(
        params.get("publisher_id") or params.get("publisher__id__exact")
    )
    if publisher_id is not None:
        qs = qs.filter(publisher__id=publisher_id)

    y_from = _parse_int(
        params.get("publish_year_from")
        or params.get("publish_year__gte")
        or params.get("publish_year")
    )
    if y_from is not None:
        qs = qs.filter(publish_year__gte=y_from)
    y_to = _parse_int(
        params.get("publish_year_to")
        or params.get("publish_year__lte")
        or params.get("publish_year")
    )
    if y_to is not None:
        qs = qs.filter(publish_year__lte=y_to)

    lang = (
        params.get("language")
        or params.get("language_code")
        or params.get("language_code__exact")
        or ""
    ).strip()
    if lang:
        qs = qs.filter(language_code__iexact=lang)

    item_status = (
        (
            params.get("item_status")
            or params.get("items__status")
            or params.get("items__status__exact")
            or ""
        )
        .strip()
        .upper()
    )
    if item_status in {"AVAILABLE", "RESERVED", "LOANED", "LOST", "DAMAGED"}:
        qs = qs.filter(items__status=item_status)

    c_from = _parse_date(
        (
            params.get("created_from")
            or params.get("created_at__date__gte")
            or ""
        ).strip()
    )
    if c_from is not None:
        qs = qs.filter(created_at__date__gte=c_from)
    c_to = _parse_date(
        (
            params.get("created_to")
            or params.get("created_at__date__lte")
            or ""
        ).strip()
    )
    if c_to is not None:
        qs = qs.filter(created_at__date__lte=c_to)

    qs = qs.distinct()

    sort = (params.get("sort") or "title").strip()
    allowed_sorts = {
        "title",
        "-title",
        "created_at",
        "-created_at",
        "publish_year",
        "-publish_year",
    }
    if sort not in allowed_sorts:
        sort = "title"
    return qs.order_by(sort)


DEFAULT_BOOK_COLUMNS: List[str] = [
    "id",
    "title",
    "authors",
    "categories",
    "isbn13",
    "publisher",
    "publisher_founded_year",
    "publish_year",
    "language",
    "cover_url",
    "total_items",
    "available_items",
    "reserved_items",
    "loaned_items",
    "lost_items",
    "damaged_items",
    "total_loans",
    "last_loan_date",
    "avg_rating",
    "total_favorites",
    "created_at",
]


def _book_cell_value(book: Book, key: str):
    if key == "id":
        return book.id
    if key == "title":
        return book.title
    if key == "authors":
        return ", ".join(a.name for a in book.authors.all())
    if key == "categories":
        return ", ".join(c.name for c in book.categories.all())
    if key == "isbn13":
        return book.isbn13 or ""
    if key == "publisher":
        return book.publisher.name if book.publisher else ""
    if key == "publisher_founded_year":
        return (
            getattr(book.publisher, "founded_year", "")
            if book.publisher
            else ""
        )
    if key == "publish_year":
        return book.publish_year or ""
    if key == "language":
        return book.language_code or ""
    if key == "cover_url":
        return book.cover_url or ""
    if key in {
        "total_items",
        "available_items",
        "reserved_items",
        "loaned_items",
        "lost_items",
        "damaged_items",
        "total_favorites",
        "authors_count",
        "categories_count",
        "total_loans",
    }:
        return getattr(book, key, 0) or 0
    if key == "avg_rating":
        val = getattr(book, "avg_rating", None)
        return round(val, 2) if val is not None else ""
    if key == "last_loan_date":
        val = getattr(book, "last_loan_date", None)
        return val.isoformat() if val else ""
    if key == "created_at":
        return (
            timezone.localtime(book.created_at).strftime("%Y-%m-%d %H:%M:%S")
            if book.created_at
            else ""
        )
    # Unknown key
    return ""


def render_books_workbook(
    qs: QuerySet[Book],
    columns: Optional[Sequence[str]] = None,
    include_items: bool = False,
) -> Workbook:
    """Render a workbook containing Books (and optionally Items) from queryset."""
    cols = list(columns or DEFAULT_BOOK_COLUMNS)
    wb = Workbook()
    ws = wb.active
    ws.title = "Books"

    # Headers
    headers = [col.replace("_", " ").title() for col in cols]
    ws.append(headers)

    for book in qs.iterator(chunk_size=500):
        row = [_book_cell_value(book, c) for c in cols]
        ws.append(row)

    # Auto-size columns
    for col_idx in range(1, len(cols) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for column_cells in ws.iter_cols(
            min_col=col_idx, max_col=col_idx, min_row=2
        ):
            for cell in column_cells:
                val = cell.value
                if val is None:
                    continue
                length = len(str(val))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max_len + 2, 60
        )

    if include_items:
        ws2 = wb.create_sheet("Items")
        item_headers = [
            "Book ID",
            "Title",
            "Barcode",
            "Status",
            "Location",
            "Created At",
            "Currently Loaned",
            "Due Date",
            "Loan Status",
        ]
        ws2.append(item_headers)
        # We rely on prefetch of items and their loans
        for book in qs.iterator(chunk_size=200):
            for item in book.items.all():
                # Determine current loan, if any
                current = next(
                    (ln for ln in item.loans.all() if ln.status == "BORROWED"),
                    None,
                )
                ws2.append(
                    [
                        book.id,
                        book.title,
                        item.barcode,
                        item.get_status_display(),
                        item.location_code or "",
                        (
                            timezone.localtime(item.created_at).strftime(
                                "%Y-%m-%d %H:%M:%S"
                            )
                            if item.created_at
                            else ""
                        ),
                        "Yes" if current else "No",
                        current.due_date.isoformat() if current else "",
                        current.get_status_display() if current else "",
                    ]
                )
        # Auto-size second sheet
        for col_idx in range(1, len(item_headers) + 1):
            max_len = len(str(item_headers[col_idx - 1]))
            for column_cells in ws2.iter_cols(
                min_col=col_idx, max_col=col_idx, min_row=2
            ):
                for cell in column_cells:
                    val = cell.value
                    if val is None:
                        continue
                    length = len(str(val))
                    if length > max_len:
                        max_len = length
            ws2.column_dimensions[get_column_letter(col_idx)].width = min(
                max_len + 2, 60
            )

    return wb
